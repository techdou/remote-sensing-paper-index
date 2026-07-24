import json
import re
import unicodedata
from pathlib import Path
from urllib.parse import quote

import openpyxl


SOURCE = Path("遥感图像分类23-26.xlsx")
COLLECTION = Path("遥感图像分类论文合集")
OUTPUT = Path("data/papers.json")


def english_chinese_title(value):
    value = str(value or "")
    en = re.search(r"\[EN\]\s*(.*?)(?=\n\s*\n\[中文\]|\n\[中文\]|$)", value, re.S)
    zh = re.search(r"\[中文\]\s*(.*)$", value, re.S)
    return (
        (en.group(1) if en else value.splitlines()[0]).strip(),
        (zh.group(1) if zh else "").strip(),
    )


def slug(value):
    normalized = unicodedata.normalize("NFKD", value)
    text = "".join(c.lower() if c.isalnum() else "-" for c in normalized)
    return re.sub(r"-+", "-", text).strip("-")[:90]


def pdf_index():
    result = {}
    for path in COLLECTION.rglob("*.pdf"):
        result.setdefault(path.stem.casefold(), path)
    return result


def find_pdf(title, index):
    clean = re.sub(r'[/:*?"<>|\\\x00-\x1f]', "_", unicodedata.normalize("NFKC", title))
    clean = re.sub(r"\s+", " ", clean).strip(" .")[:180].rstrip(" .")
    path = index.get(clean.casefold())
    if not path:
        return ""
    return "./" + "/".join(quote(part) for part in path.parts)


def main():
    workbook = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    pdfs = pdf_index()
    papers = []
    seen_ids = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        headers = [str(cell.value or "").strip() for cell in sheet[3]]
        positions = {header: index for index, header in enumerate(headers)}
        title_index = next(i for i, value in enumerate(headers) if "Title" in value)
        venue_index = next(i for i, value in enumerate(headers) if "会议/期刊" in value)
        url_index = next(i for i, value in enumerate(headers) if "URL" in value)
        for row in sheet.iter_rows(min_row=4, values_only=True):
            if not row[title_index]:
                continue
            title_en, title_zh = english_chinese_title(row[title_index])
            base_id = slug(title_en) or f"paper-{len(papers)+1}"
            seen_ids[base_id] = seen_ids.get(base_id, 0) + 1
            paper_id = base_id if seen_ids[base_id] == 1 else f"{base_id}-{seen_ids[base_id]}"
            value = lambda label, fallback="": str(row[positions[label]] or "") if label in positions else fallback
            papers.append({
                "id": paper_id,
                "sheet": sheet_name,
                "titleEn": title_en,
                "titleZh": title_zh,
                "abstract": value("Abstract"),
                "year": row[next(i for i, h in enumerate(headers) if "年" in h and ("正式" in h or h == "年份"))] or "",
                "venue": str(row[venue_index] or "未分类"),
                "relation": value("遥感分类关系", value("研究")),
                "positioning": value("分类定位（任务+标签形式）", value("层级")),
                "modalities": value("输入模态"),
                "mechanism": value("核心机制"),
                "evaluation": value("遥感数据集/评估"),
                "sourceUrl": str(row[url_index] or ""),
                "pdfUrl": find_pdf(title_en, pdfs),
            })
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(papers, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"records": len(papers), "withPdf": sum(bool(p["pdfUrl"]) for p in papers)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
